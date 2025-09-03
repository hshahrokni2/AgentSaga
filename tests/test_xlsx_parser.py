"""
Test suite for XLSX parser with Swedish character support.
Following TDD principles - these tests are written before implementation.
"""
import unittest
from unittest.mock import Mock, patch, MagicMock
import pandas as pd
from datetime import datetime
from pathlib import Path
import tempfile
import openpyxl

# Import the modules we're testing (will fail initially as they don't exist)
from src.parsers.xlsx_parser import SwedishXLSXParser
from src.parsers.exceptions import (
    InvalidFileFormatError,
    CharacterEncodingError,
    CorruptedFileError
)


class TestSwedishXLSXParser(unittest.TestCase):
    """Test XLSX parsing with Swedish characters (åäö/ÅÄÖ)"""
    
    def setUp(self):
        """Set up test fixtures"""
        self.parser = SwedishXLSXParser()
        self.test_data_dir = Path(tempfile.mkdtemp())
        
    def tearDown(self):
        """Clean up test fixtures"""
        import shutil
        shutil.rmtree(self.test_data_dir, ignore_errors=True)
        
    def test_parse_xlsx_with_swedish_characters_in_headers(self):
        """Test that parser correctly handles Swedish characters in column headers"""
        # Arrange
        test_file = self.test_data_dir / "test_åäö.xlsx"
        test_data = {
            'Återvinning': ['Plast', 'Glas', 'Metall'],
            'Mängd (kg)': [100.5, 200.75, 50.25],
            'Insamlingsområde': ['Södertälje', 'Västerås', 'Örebro']
        }
        df = pd.DataFrame(test_data)
        df.to_excel(test_file, index=False)
        
        # Act
        result = self.parser.parse(test_file)
        
        # Assert
        self.assertIsNotNone(result)
        self.assertIn('Återvinning', result.columns)
        self.assertIn('Mängd (kg)', result.columns)
        self.assertIn('Insamlingsområde', result.columns)
        self.assertEqual(len(result), 3)
        self.assertEqual(result['Insamlingsområde'].iloc[0], 'Södertälje')
        
    def test_parse_xlsx_with_swedish_characters_in_data(self):
        """Test that parser correctly handles Swedish characters in cell data"""
        # Arrange
        test_file = self.test_data_dir / "test_data.xlsx"
        test_data = {
            'Material': ['Återvunnet glas', 'Förpackningar', 'Farligt avfall'],
            'Kategori': ['Återvinning', 'Återanvändning', 'Specialbehandling'],
            'Ansvarig': ['Åsa Andersson', 'Örjan Öberg', 'Ärling Ek']
        }
        df = pd.DataFrame(test_data)
        df.to_excel(test_file, index=False)
        
        # Act
        result = self.parser.parse(test_file)
        
        # Assert
        self.assertEqual(result['Material'].iloc[0], 'Återvunnet glas')
        self.assertEqual(result['Ansvarig'].iloc[1], 'Örjan Öberg')
        self.assertTrue(all(isinstance(val, str) for val in result['Ansvarig']))
        
    def test_parse_xlsx_with_mixed_encoding(self):
        """Test handling of files with mixed character encodings"""
        # Arrange
        test_file = self.test_data_dir / "mixed_encoding.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Återvinning'
        ws['B1'] = 'Mängd'
        ws['A2'] = 'Plåtburkar'  # Swedish å
        ws['B2'] = '100,5'  # Swedish decimal comma
        wb.save(test_file)
        
        # Act
        result = self.parser.parse(test_file, encoding='utf-8')
        
        # Assert
        self.assertEqual(result.iloc[0]['Återvinning'], 'Plåtburkar')
        
    def test_parse_xlsx_invalid_file_format(self):
        """Test that invalid file formats raise appropriate errors"""
        # Arrange
        test_file = self.test_data_dir / "invalid.txt"
        test_file.write_text("This is not an Excel file")
        
        # Act & Assert
        with self.assertRaises(InvalidFileFormatError) as ctx:
            self.parser.parse(test_file)
        self.assertIn("not a valid XLSX file", str(ctx.exception))
        
    def test_parse_xlsx_corrupted_file(self):
        """Test handling of corrupted XLSX files"""
        # Arrange
        test_file = self.test_data_dir / "corrupted.xlsx"
        test_file.write_bytes(b"PK\x03\x04corrupted_data")  # Partial XLSX header
        
        # Act & Assert
        with self.assertRaises(CorruptedFileError) as ctx:
            self.parser.parse(test_file)
        self.assertIn("corrupted or unreadable", str(ctx.exception))
        
    def test_parse_xlsx_empty_file(self):
        """Test handling of empty XLSX files"""
        # Arrange
        test_file = self.test_data_dir / "empty.xlsx"
        wb = openpyxl.Workbook()
        wb.save(test_file)
        
        # Act
        result = self.parser.parse(test_file, allow_empty=True)
        
        # Assert
        self.assertTrue(result.empty)
        
    def test_parse_xlsx_large_file_streaming(self):
        """Test streaming parse for large files"""
        # Arrange
        test_file = self.test_data_dir / "large.xlsx"
        large_data = {
            'ID': range(10000),
            'Återvinningstyp': ['Plast' if i % 2 == 0 else 'Glas' for i in range(10000)],
            'Vikt': [100.5 + i for i in range(10000)]
        }
        df = pd.DataFrame(large_data)
        df.to_excel(test_file, index=False)
        
        # Act
        chunk_count = 0
        for chunk in self.parser.parse_stream(test_file, chunksize=1000):
            chunk_count += 1
            self.assertLessEqual(len(chunk), 1000)
            
        # Assert
        self.assertEqual(chunk_count, 10)
        
    def test_parse_xlsx_multiple_sheets(self):
        """Test parsing XLSX with multiple sheets"""
        # Arrange
        test_file = self.test_data_dir / "multi_sheet.xlsx"
        with pd.ExcelWriter(test_file) as writer:
            df1 = pd.DataFrame({'Återvinning': ['Plast'], 'Mängd': [100]})
            df2 = pd.DataFrame({'Förbränning': ['Trä'], 'Volym': [50]})
            df1.to_excel(writer, sheet_name='Återvinning', index=False)
            df2.to_excel(writer, sheet_name='Förbränning', index=False)
            
        # Act
        result = self.parser.parse_all_sheets(test_file)
        
        # Assert
        self.assertIn('Återvinning', result)
        self.assertIn('Förbränning', result)
        self.assertEqual(result['Återvinning']['Mängd'].iloc[0], 100)
        
    def test_parse_xlsx_with_formulas(self):
        """Test handling of XLSX files with formulas"""
        # Arrange
        test_file = self.test_data_dir / "formulas.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Återvinning'
        ws['B1'] = 'Vikt'
        ws['C1'] = 'Total'
        ws['A2'] = 'Plast'
        ws['B2'] = 100
        ws['C2'] = '=B2*1.25'  # Formula
        wb.save(test_file)
        
        # Act
        result = self.parser.parse(test_file, evaluate_formulas=True)
        
        # Assert
        self.assertEqual(result['Total'].iloc[0], 125.0)


class TestXLSXParserPerformance(unittest.TestCase):
    """Performance tests for XLSX parser"""
    
    def setUp(self):
        self.parser = SwedishXLSXParser()
        self.test_data_dir = Path(tempfile.mkdtemp())
        
    def tearDown(self):
        import shutil
        shutil.rmtree(self.test_data_dir, ignore_errors=True)
        
    def test_parse_performance_small_file(self):
        """Test that small files parse in under 100ms"""
        # Arrange
        test_file = self.test_data_dir / "small.xlsx"
        df = pd.DataFrame({'A': range(10), 'B': range(10)})
        df.to_excel(test_file, index=False)
        
        # Act
        import time
        start = time.perf_counter()
        result = self.parser.parse(test_file)
        duration = time.perf_counter() - start
        
        # Assert
        self.assertLess(duration, 0.1)  # Should parse in under 100ms
        self.assertEqual(len(result), 10)
        
    def test_memory_efficient_parsing(self):
        """Test memory-efficient parsing for large files"""
        # Arrange
        test_file = self.test_data_dir / "memory_test.xlsx"
        large_data = {'Col' + str(i): range(1000) for i in range(50)}
        df = pd.DataFrame(large_data)
        df.to_excel(test_file, index=False)
        
        # Act
        import tracemalloc
        tracemalloc.start()
        result = self.parser.parse(test_file, low_memory=True)
        current, peak = tracemalloc.get_traced_memory()
        tracemalloc.stop()
        
        # Assert
        self.assertLess(peak / 1024 / 1024, 100)  # Should use less than 100MB
        

if __name__ == '__main__':
    unittest.main()